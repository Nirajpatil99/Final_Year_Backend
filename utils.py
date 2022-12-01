from openpyxl.workbook import Workbook as openpyxlWorkbook
import pandas as pd
from openpyxl import load_workbook
import xlrd
import json
import os
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

frontEndURL = "http://localhost:3000"


def convertToXLSX(orgFileName, newFileName):
    xlsBook = xlrd.open_workbook(orgFileName)
    workbook = openpyxlWorkbook()
    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(4, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row - 3, column=col +
                           1).value = xlsSheet.cell_value(row, col)
    workbook.save(newFileName)
    workbook = load_workbook(newFileName)
    return workbook


def generateJsonForDailyProd(workbook):
    res = {}
    for ws in workbook.worksheets:
        prev = ""
        jobOrder = []
        partName = []
        operation = []
        cy_time = []
        part_running = []
        operator_name = []
        actual_prod_time = []
        prod_per_hr = []
        qty_req = []
        qty_achieved = []
        loss_qty = []
        loss_mc_hr = []
        st = []
        programmer_busy = []

        for idx, i in enumerate(ws.rows):
            if (idx == 0):
                continue
            if (i[0].value != None):
                prev = i[0].value
            else:
                i[0].value = prev
            if (i[1].value == None):
                break
            jobOrder.append(i[1].value)
            partName.append(i[2].value)
            operation.append(i[3].value)
            cy_time.append(i[4].value)
            part_running.append(i[5].value)
            operator_name.append(i[6].value)
            actual_prod_time.append(i[7].value)
            prod_per_hr.append(i[8].value)
            qty_req.append(i[9].value)
            qty_achieved.append(i[10].value)
            loss_qty.append(i[11].value)
            loss_mc_hr.append(i[12].value)
            st.append(i[13].value)
            programmer_busy.append(i[14].value)

        combined = {"joborders": jobOrder, "partname": partName,
                    "operation": operation, "cy_time": cy_time, "part_running": part_running, "operator_name": operator_name,
                    "actual_prod_time": actual_prod_time, "prod_per_hr": prod_per_hr, "qty_req": qty_req, "qty_achieved": qty_achieved,
                    "loss_qty": loss_qty, "loss_mc_hr": loss_mc_hr, "st": st, "programmer_busy": programmer_busy}

        df = pd.DataFrame.from_dict(combined)

        # print(df[df.qty_achieved == df.qty_achieved.max()]["qty_achieved"])
        # print(df.dtypes)

        df['actual_prod_time'] = df['actual_prod_time'].astype(float).round(2)
        df['prod_per_hr'] = df['prod_per_hr'].astype(float).round(2)
        df['qty_req'] = df['qty_req'].astype(float).round(2)
        df['loss_qty'] = df['loss_qty'].astype(float).round(2)
        df['loss_mc_hr'] = df['loss_mc_hr'].astype(float).round(2)
        df['st'] = df['st'].astype(float).round(2)
        df['programmer_busy'] = df['programmer_busy'].astype(float).round(2)
        df['qty_achieved'] = df['qty_achieved'].astype(float).round(2)

        # df.round(2)

        # df.to_json("test.json", orient="records")
        partGrp = df.groupby(['partname'])
        res[ws.title] = {}
        rows = []
        for key in list(partGrp.groups.keys()):
            row = partGrp.get_group(
                key).groupby(['partname']).sum().to_dict(orient="records")[0]
            row['partname'] = key
            rows.append(row)

        majors = []

        majors.append({"max_lossqty": df[["partname", "loss_qty"]].sort_values(
            by="loss_qty", ascending=False).head(5).to_dict(orient="records")})

        majors.append({"max_actual_prod_time": df[["partname", "actual_prod_time"]].sort_values(
            by="actual_prod_time", ascending=False).head(5).to_dict(orient="records")
        })
        majors.append({"max_prod_per_hr": df[["partname", "prod_per_hr"]].sort_values(
            by="prod_per_hr", ascending=False).head(5).to_dict(orient="records")})

        majors.append({"least_prod_per_hr": df[["partname", "prod_per_hr"]].sort_values(
            by="prod_per_hr", ascending=True).head(5).to_dict(orient="records")})

        majors.append({"max_loss_mc_hr": df[["partname", "loss_mc_hr"]].sort_values(
            by="loss_mc_hr", ascending=False).head(5).to_dict(orient="records")})

        majors.append({"least_loss_mc_hr": df[["partname", "loss_mc_hr"]].sort_values(
            by="loss_mc_hr", ascending=True).head(5).to_dict(orient="records")})

        majors.append({"max_qty_req": df[["partname", "qty_req"]].sort_values(
            by="qty_req", ascending=False).head(5).to_dict(orient="records")})

        majors.append({"max_qty_achieved": df[["partname", "qty_achieved"]].sort_values(
            by="qty_achieved", ascending=False).head(5).to_dict(orient="records")})

        majorTitles = ["Top 5 Parts with Maximum Loss",
                       "Top 5 Parts with Actual Production", "Top 5 Parts with Maximum prod per hour", "Top 5 parts with Least Production Per Hour", "Top 5 parts with Maximum MC hours loss", "Top 5 parts with Least MC hours loss", "Top 5 Parts with Maximum Quantity Required", "Top 5 Parts with Maximum Quantity Achieved"]

        res[ws.title].update(
            {"rows": rows, "majors": majors, "majorTitles": majorTitles})

        # print(res)
        # res = res

    # print(json.dumps(res))
    return res


def handleDailyProd(filename):
    workBook = convertToXLSX(filename, filename.replace("xls", "xlsx"))
    # print(workBook.worksheets)
    # return {}
    return generateJsonForDailyProd(workBook)


def handleMisProd(filename):
    workBook = load_workbook(filename, data_only=True)
    return {'FC - Operations - In House Prod': handleMisProdInHouseProd(
        workBook['FC - Operations - In House Prod']), "FC - Machine Shop": handleMisProdMachineShop(workBook["FC - Machine Shop"])}


def handleMisProdMachineShop(ws):
    dic = {}
    mainH = ws.cell(10, 2).value
    dic[mainH] = []
    # row = Row , col = Col
    for row in range(6, 1000, 5):
        datee = ws.cell(4, row)
        if (datee.value == None):
            break
        datee = datee.value.strftime("%b-%Y")
        dateDic = {}
        for col in range(11, 16):
            key = ws.cell(col, 2).value.strip()
            dateDic[key] = []
            for k in range(0, 4):
                valToAppend = ws.cell(col, row+k).value
                dateDic[key].append(
                    valToAppend if valToAppend != None else 0)
        dic[mainH].append(dateDic)

    mainH = ws.cell(17, 2).value
    dic[mainH] = []
    dic["months"] = []
    # row = Row , col = Col
    for row in range(6, 1000, 5):
        datee = ws.cell(4, row)
        if (datee.value == None):
            break
        datee = datee.value.strftime("%b-%Y")
        dic["months"].append(datee)
        dateDic = {}
        for col in range(18, 23):
            key = ws.cell(col, 2).value.strip()
            dateDic[key] = []
            for k in range(0, 4):
                valToAppend = ws.cell(col, row+k).value
                dateDic[key].append(
                    valToAppend if valToAppend != None else 0)
        dic[mainH].append(dateDic)

    # lol = []
    # for key, value in dic.items():
    #     # temp = dict()
    #     # temp['detail'] = key
    #     for key1, value1 in value.items():
    #         # temp['month'] = key1
    #         for key2, value2 in value1.items():
    #             # print(value2[0])
    #             # temp['subdetail'] = key2
    #             # temp['fitting'] = value2[0]
    #             # temp['valves'] = value2[1]
    #             # temp['clamp'] = value2[2]
    #             lol.append({'detail': key, 'month': key1,
    #                         'subdetail': key2, 'fitting': value2[0], 'valves': value2[1], 'clamp': value2[2]})
    #     temp = dict()

    # print(pd.DataFrame.from_records(lol).to_json(orient="records"))
    return dic


def handleMisProdInHouseProd(ws):
    dic = {}
    mainH = ws.cell(8, 2).value
    dic[mainH] = []
    # row = Row , col = Col
    for row in range(6, 1000, 5):
        datee = ws.cell(4, row)
        if (datee.value == None):
            break
        datee = datee.value.strftime("%b-%Y")
        dateDic = {}

        for col in range(9, 14):
            dateDic[ws.cell(col, 2).value] = []
            for k in range(0, 4):
                valToAppend = ws.cell(col, row+k).value
                dateDic[ws.cell(col, 2).value].append(
                    valToAppend if valToAppend != None else 0)
        dic[mainH].append(dateDic)

    mainH = ws.cell(15, 2).value
    dic[mainH] = []
    # row = Row , col = Col
    for row in range(6, 1000, 5):
        datee = ws.cell(4, row)
        if (datee.value == None):
            break
        datee = datee.value.strftime("%b-%Y")
        dateDic = {}
        for col in range(17, 20):
            dateDic[ws.cell(col, 2).value] = []
            for k in range(0, 4):
                valToAppend = ws.cell(col, row+k).value
                dateDic[ws.cell(col, 2).value].append(
                    valToAppend if valToAppend != None else 0)
        dic[mainH].append(dateDic)

    mainH = ws.cell(21, 2).value
    dic[mainH] = []
    # row = Row , col = Col
    for row in range(6, 1000, 5):
        datee = ws.cell(4, row)
        if (datee.value == None):
            break
        datee = datee.value.strftime("%b-%Y")
        dateDic = {}
        for col in range(23, 29):
            dateDic[ws.cell(col, 2).value] = []
            for k in range(0, 4):
                valToAppend = ws.cell(col, row+k).value
                dateDic[ws.cell(col, 2).value].append(
                    valToAppend if valToAppend != None else 0)
        dic[mainH].append(dateDic)

    mainH = ws.cell(30, 2).value
    dic[mainH] = []
    dic["months"] = []
    # row = Row , col = Col
    for row in range(6, 1000, 5):
        datee = ws.cell(4, row)
        if (datee.value == None):
            break
        datee = datee.value.strftime("%b-%Y")
        dic["months"].append(datee)
        dateDic = {}
        for col in range(33, 36):
            dateDic[ws.cell(col, 3).value] = []
            for k in range(0, 4):
                valToAppend = ws.cell(col, row+k).value
                dateDic[ws.cell(col, 3).value].append(
                    valToAppend if valToAppend != None else 0)
        dic[mainH].append(dateDic)

    # lol = []
    # for key, value in dic.items():
    #     # temp = dict()
    #     # temp['detail'] = key
    #     for key1, value1 in value.items():
    #         # temp['month'] = key1
    #         for key2, value2 in value1.items():
    #             # print(value2[0])
    #             # temp['subdetail'] = key2
    #             # temp['fitting'] = value2[0]
    #             # temp['valves'] = value2[1]
    #             # temp['clamp'] = value2[2]
    #             lol.append({'detail': key, 'month': key1,
    #                         'subdetail': key2, 'fitting': value2[0], 'valves': value2[1], 'clamp': value2[2]})
    return dic


def deleteConvertedXLS(filename):
    if (os.path.exists(filename)):
        os.remove(filename)
    filename = filename.replace("xls", "xlsx")
    if (os.path.exists(filename)):
        os.remove(filename)
    pass


def sendForgotPasswordMail(receiver_email, name, uniqueID):
    port = 465  # For SSL
    password = ""

    # Create a secure SSL context
    context = ssl.create_default_context()

    sender_email = "tejasvp252@gmail.com"
    # receiver_email = "tejasvp25@gmail.com"
    # name = "Tejas"
    link = "{}/resetpassword/{}".format(frontEndURL, uniqueID)

    mailHTML = """\
        <table width="95%" border="0" align="center" cellpadding="0" cellspacing="0"
    style="max-width:670px;background:#fff; border-radius:3px; text-align:center;-webkit-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);-moz-box-shadow:0 6px 18px 0 rgba(0,0,0,.06);box-shadow:0 6px 18px 0 rgba(0,0,0,.06);">
    <tr>
        <td style="height:40px;">&nbsp;</td>
    </tr>
    <tr>
        <td style="padding:0 35px;">
            <h1 style="color:#1e1e2d; font-weight:500; margin:0;font-size:32px;font-family:'Rubik',sans-serif;">Hello {}</h1>
            <br>
            <h2 style="color:#1e1e2d; font-weight:500; margin:0;font-family:'Rubik',sans-serif;">You have requested to reset your password</h2>
            <span
                style="display:inline-block; vertical-align:middle; margin:29px 0 26px; border-bottom:1px solid #cecece; width:100px;"></span>
            <p style="color:#455056; font-size:15px;line-height:24px; margin:0;font-family:'Rubik',sans-serif;">A unique link to reset your
                password has been generated for you. To reset your password, click the
                following button and follow the instructions.
            </p>
            <a href={}
                style="background:#20e277;text-decoration:none !important; font-weight:500; margin-top:35px; color:#fff;text-transform:uppercase; font-size:14px;padding:10px 24px;display:inline-block;border-radius:50px;font-family:'Rubik',sans-serif;">Reset
                Password</a>
        </td>
    </tr>
    <tr>
        <td style="height:40px;">&nbsp;</td>
    </tr>
    </table>""".format(name, link)

    message = MIMEMultipart()
    message["Subject"] = "Forgot Password"
    message["From"] = sender_email
    message["To"] = receiver_email
    message.attach(MIMEText(mailHTML, "html"))

    server = smtplib.SMTP_SSL("smtp.gmail.com", port, context=context)
    try:
        server.login("tejasvp252@gmail.com", password)
        server.sendmail(sender_email, receiver_email, message.as_string())
    except Exception:
        print(Exception)
        # raise Exception
        return False
    finally:
        server.quit()
    return True
